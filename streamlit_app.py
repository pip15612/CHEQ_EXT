import streamlit as st
import cv2
import numpy as np
import pytesseract
import easyocr
from pdf2image import convert_from_bytes
from PIL import Image
import pandas as pd
import openpyxl
import re
from io import BytesIO
import os
import traceback
import shutil
from pathlib import Path
from datetime import datetime

# ========== Page Config ==========
st.set_page_config(page_title="Thai Cheque OCR", page_icon="🏦", layout="wide")

# ========== Setup Tessdata for MICR ==========
@st.cache_resource
def setup_tessdata():
    """จัดการ e13b.traineddata สำหรับ MICR"""
    try:
        tessdata_dir = Path("/tmp/tessdata")
        tessdata_dir.mkdir(parents=True, exist_ok=True)
        
        local_e13b = Path("e13b.traineddata")
        target_e13b = tessdata_dir / "e13b.traineddata"
        
        if local_e13b.exists() and not target_e13b.exists():
            shutil.copy(local_e13b, target_e13b)
        
        return str(tessdata_dir)
    except Exception as e:
        st.warning(f"⚠️ ไม่สามารถตั้งค่า e13b: {str(e)}")
        return None

CUSTOM_TESSDATA = setup_tessdata()

# ========== Initialize EasyOCR ==========
@st.cache_resource
def initialize_easyocr():
    """Initialize EasyOCR reader (cached)"""
    with st.spinner("🔄 กำลังโหลด EasyOCR models... (ครั้งแรกใช้เวลา 2-3 นาที)"):
        return easyocr.Reader(['th', 'en'], gpu=False)

# ========== Helper Functions ==========
def _is_template_date_line(text: str) -> bool:
    """กันบรรทัดแม่แบบวันที่"""
    t = (text or "").strip().lower()
    template_words = ["day", "month", "year", "dd", "mm", "yyyy", "วว", "ดด", "ปปปป", "วัน", "เดือน", "ปี"]
    hits = sum(1 for w in template_words if w in t)
    digit_count = len(re.findall(r"\d", t))
    return (hits >= 2 and digit_count <= 4)

def _validate_date(d: str, m: str, y: str):
    """Validate และคืน dd/mm/yyyy"""
    try:
        di, mi, yi = int(d), int(m), int(y)
        yi_check = yi - 543 if yi > 2400 else yi
        if 1 <= di <= 31 and 1 <= mi <= 12 and 1990 <= yi_check <= 2040:
            return f"{d}/{m}/{y}"
    except:
        return ""
    return ""

def clean_messy_date(text):
    """Robust date parser with sliding window"""
    if not text:
        return ""
    
    text = re.sub(r'(?i)(วันที่|วันที|date|of)\s*[:\-]?\s*', '', text).strip()
    if _is_template_date_line(text):
        return ""
    
    # (1) dd/mm/yyyy with separators
    m = re.search(r'(\d{1,2})\s*[\/\-\.\s]\s*(\d{1,2})\s*[\/\-\.\s]\s*(\d{2,4})', text)
    if m:
        d, mo, y = m.group(1), m.group(2), m.group(3)
        if len(d) == 1: d = "0" + d
        if len(mo) == 1: mo = "0" + mo
        if len(y) == 2: y = "20" + y
        out = _validate_date(d, mo, y)
        if out:
            return out
    
    # (2) Sliding 8-digit window
    digits = "".join([c for c in text if c.isdigit()])
    if len(digits) >= 8:
        for start in range(0, len(digits) - 8 + 1):
            w = digits[start:start+8]
            d, mo, y = w[:2], w[2:4], w[4:]
            out = _validate_date(d, mo, y)
            if out:
                return out
    
    return ""

def extract_cheque_digit(micr_text):
    if not micr_text:
        return ""
    digits = "".join(c for c in micr_text if c.isdigit())
    return digits[:2] if len(digits) >= 2 else ""

def parse_micr_thai(micr_text):
    parts = [re.sub(r'[^\d]', '', p.strip())
             for p in re.split(r'[⑆⑇⑈⑉]', micr_text)
             if p.strip()]
    
    chq_no = bank_cd = branch_cd = acc_no = ""
    if len(parts) < 2:
        return chq_no, bank_cd, branch_cd, acc_no
    
    chq_no = parts[1]
    
    if len(parts) >= 5 and len(parts[2]) == 3:
        bank_cd = parts[2]
        branch_cd = parts[3]
        acc_no = parts[4]
        return chq_no, bank_cd, branch_cd, acc_no
    
    if len(parts) >= 3:
        raw = parts[2]
        if len(raw) >= 3:
            bank_cd = raw[:3]
        if len(raw) >= 7:
            branch_cd = raw[3:7]
            acc_no = raw[7:] if len(raw) > 7 else ""
        elif len(raw) > 3:
            branch_cd = raw[3:]
    
    if not acc_no and len(parts) >= 4:
        acc_no = parts[3]
    
    return chq_no, bank_cd, branch_cd, acc_no

def clean_amount_garbage(text):
    text = re.sub(r'(?i)(baht|bath|amount|จ่าย|pay|[^ก-๙\s])', '', text)
    return text.replace("*", "").replace("=", "").strip()

def clean_payee_final(text):
    typos = {"บรบท": "บริษัท", "บริบัท": "บริษัท", "จากัด": "จำกัด"}
    for w, c in typos.items():
        text = text.replace(w, c)
    return re.sub(r'(หรือผู้ถือ|Or Bearer).*$', '', text, flags=re.IGNORECASE).strip(" .-_^$#/")

def robust_auto_crop(image):
    img_h, img_w = image.shape[:2]
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(gray, (7, 7), 0)
    edged = cv2.Canny(blur, 50, 150)
    kernel = np.ones((5, 5), np.uint8)
    dilated = cv2.dilate(edged, kernel, iterations=2)
    contours, _ = cv2.findContours(dilated, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    contours = sorted(contours, key=cv2.contourArea, reverse=True)[:5]
    
    for c in contours:
        peri = cv2.arcLength(c, True)
        approx = cv2.approxPolyDP(c, 0.02 * peri, True)
        if len(approx) == 4 and cv2.contourArea(c) > (img_w * img_h * 0.35):
            x, y, w, h = cv2.boundingRect(approx)
            return image[max(0, y-20):min(img_h, y+h+20), max(0, x-60):min(img_w, x+w+60)]
    
    return image[int(img_h*0.15):int(img_h*0.85), int(img_w*0.02):int(img_w*0.98)]

def crop_micr_region(image_bgr):
    h, w = image_bgr.shape[:2]
    return image_bgr[int(h * 0.78):h, 0:w]

def extract_micr(image):
    crop = crop_micr_region(image)
    gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
    _, thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    
    micr_text = ""
    if CUSTOM_TESSDATA:
        try:
            micr_text = pytesseract.image_to_string(
                thresh, 
                config=f'--psm 7 --tessdata-dir {CUSTOM_TESSDATA} -l e13b'
            ).strip()
        except:
            pass
    
    if not micr_text:
        try:
            micr_text = pytesseract.image_to_string(thresh, lang='eng', config='--psm 7').strip()
        except:
            return ""
    
    mapping = {'A': '⑆', 'B': '⑇', 'C': '⑈', 'D': '⑉'}
    for k, v in mapping.items():
        micr_text = micr_text.replace(k, v).replace(k.lower(), v)
    
    return "".join([c for c in micr_text if c in "0123456789⑆⑇⑈⑉ "])

def group_text_into_lines(ocr_results):
    lines = []
    sorted_res = sorted(ocr_results, key=lambda r: r[0][0][1])
    
    for bbox, text, conf in sorted_res:
        matched = False
        for line in lines:
            y_min = min([item[0][0][1] for item in line])
            y_max = max([item[0][2][1] for item in line])
            curr_min, curr_max = bbox[0][1], bbox[2][1]
            inter = min(y_max, curr_max) - max(y_min, curr_min)
            if inter > 0 and (inter / min(y_max - y_min, curr_max - curr_min)) > 0.5:
                line.append((bbox, text))
                matched = True
                break
        if not matched:
            lines.append([(bbox, text)])
    
    for line in lines:
        line.sort(key=lambda item: item[0][0][0])
    
    return lines

def extract_thai_data(image, reader):
    """ดึงข้อมูลหลักจากเช็ค"""
    img_h, img_w, _ = image.shape
    
    raw_results = reader.readtext(image, detail=1, paragraph=False)
    lines = group_text_into_lines(raw_results)
    
    data = {"Date": "", "Payee": "", "Amount_Text": "", "Amount_Num": ""}
    money_kws = ["บาท", "Baht", "ถ้วน", "ล้าน", "แสน", "หมื่น", "พัน", "ร้อย", "สิบ"]
    
    for i, line in enumerate(lines):
        full_line_text = " ".join([item[1] for item in line]).strip()
        
        # Date extraction
        date_kw_hit = (
            ("วันที่" in full_line_text) or 
            ("วันที" in full_line_text) or
            (re.search(r'(?i)\bdate\b', full_line_text) is not None)
        )
        
        if date_kw_hit and not data["Date"]:
            if not _is_template_date_line(full_line_text):
                d0 = clean_messy_date(full_line_text)
                if d0:
                    data["Date"] = d0
        
        # Amount text
        if any(k in full_line_text for k in money_kws) and re.search(r'[ก-๙]', full_line_text):
            cleaned = clean_amount_garbage(full_line_text)
            if len(cleaned) > len(data["Amount_Text"]):
                data["Amount_Text"] = cleaned
        
        # Payee
        pay_kws = ["จ่าย", "Pay", "แก่", "to"]
        if any(kw in full_line_text for kw in pay_kws) and not any(k in full_line_text for k in money_kws):
            name = full_line_text
            for k in pay_kws:
                name = name.replace(k, "")
            name = name.split("วันที่")[0].strip(" .-_/^*")
            if len(name) > 2 and not data["Payee"]:
                data["Payee"] = clean_payee_final(name)
    
    # Amount number
    for line in lines:
        for bbox, text in line:
            if (bbox[0][0] + bbox[1][0]) / 2 > img_w * 0.5:
                money_pattern = r'\d{1,3}(?:,\d{3})*\.\d{2}'
                clean_t = text.replace(" ", "").replace("b", "").replace("B", "").replace("฿", "")
                matches = re.findall(money_pattern, clean_t)
                
                if matches:
                    candidate = max(matches, key=len)
                    if len(candidate) >= len(data["Amount_Num"]):
                        data["Amount_Num"] = candidate
    
    return data

# ========== Main Processing Function (FIX: This was missing!) ==========
def process_cheque(uploaded_file):
    """ประมวลผลไฟล์เช็ค PDF หรือรูปภาพ"""
    all_results = []
    all_cropped = []
    
    try:
        # Initialize EasyOCR
        reader = initialize_easyocr()
        
        # Convert to images
        uploaded_file.seek(0)
        if uploaded_file.name.lower().endswith('.pdf'):
            images = convert_from_bytes(uploaded_file.read(), dpi=200)
        else:
            pil_image = Image.open(uploaded_file)
            images = [pil_image]
        
        # Process each page/image
        for page_num, pil_img in enumerate(images, start=1):
            st.info(f"🔍 กำลังประมวลผลหน้า {page_num}/{len(images)}")
            
            # Convert to OpenCV format
            img_array = np.array(pil_img)
            if len(img_array.shape) == 2:
                image_bgr = cv2.cvtColor(img_array, cv2.COLOR_GRAY2BGR)
            elif img_array.shape[2] == 4:
                image_bgr = cv2.cvtColor(img_array, cv2.COLOR_RGBA2BGR)
            else:
                image_bgr = cv2.cvtColor(img_array, cv2.COLOR_RGB2BGR)
            
            # Auto crop
            cropped = robust_auto_crop(image_bgr)
            all_cropped.append(cropped)
            
            # Extract MICR
            micr_text = extract_micr(cropped)
            chq_no, bank_cd, branch_cd, acc_no = parse_micr_thai(micr_text)
            chq_digit = extract_cheque_digit(micr_text)
            
            # Extract Thai data
            thai_data = extract_thai_data(cropped, reader)
            
            # Combine results
            result = {
                "หน้า": page_num,
                "วันที่": thai_data["Date"],
                "ผู้รับเงิน": thai_data["Payee"],
                "จำนวนเงิน (ตัวอักษร)": thai_data["Amount_Text"],
                "จำนวนเงิน": thai_data["Amount_Num"],
                "หมายเลขเช็ค": chq_no,
                "Cheque digit": chq_digit,
                "รหัสธนาคาร": bank_cd,
                "รหัสสาขา": branch_cd,
                "เลขบัญชี": acc_no,
                "MICR (Raw)": micr_text
            }
            
            all_results.append(result)
        
        return all_results, all_cropped
        
    except Exception as e:
        st.error(f"❌ เกิดข้อผิดพลาดในการประมวลผล: {str(e)}")
        st.code(traceback.format_exc())
        return [], []

# ========== XLOOKUP Function ==========
def xlookup(lookup_value, lookup_array, return_array, if_not_found=None):
    """Mimics Excel's XLOOKUP function with automatic type conversion"""
    try:
        if not isinstance(lookup_array, pd.Series):
            lookup_array = pd.Series(lookup_array)
        if not isinstance(return_array, pd.Series):
            return_array = pd.Series(return_array)
        
        if isinstance(lookup_value, (int, float)):
            try:
                lookup_array_numeric = pd.to_numeric(lookup_array, errors='coerce')
                mask = lookup_array_numeric == lookup_value
            except:
                mask = lookup_array == lookup_value
        else:
            mask = lookup_array == lookup_value
            if not mask.any() and isinstance(lookup_value, str) and lookup_value.replace('.','').replace('-','').isdigit():
                try:
                    lookup_array_numeric = pd.to_numeric(lookup_array, errors='coerce')
                    lookup_value_numeric = pd.to_numeric(lookup_value, errors='coerce')
                    mask = lookup_array_numeric == lookup_value_numeric
                except:
                    pass
        
        if mask.any():
            idx = mask.idxmax()
            return return_array.iloc[idx]
        else:
            return if_not_found
    except:
        return if_not_found

# ========== Process Template Filling ==========
def process_template_filling(pdf_file, fchn_file, master_file, template_file, business_partner=""):
    """ประมวลผล Template Filling - เขียนทับแถวใหม่ และลบแถวส่วนเกินทิ้ง"""
    try:
        # Load Template
        template_wb = openpyxl.load_workbook(template_file)
        template_sheet = template_wb['TEMPLATE (TR Teams) ']
        cash_sheet = template_wb['TEMPLATE (Cash Teams)']
        
        # Load data files
        pdf_df = pd.read_excel(pdf_file, sheet_name=0, dtype=str)
        
        # Map column names
        column_mapping = {
            'เลขบัญชี': 'เลขบัญชี', 'Account number': 'เลขบัญชี',
            'หมายเลขเช็ค': 'หมายเลขเช็ค', 'Cheque Number': 'หมายเลขเช็ค',
            'จำนวนเงิน': 'จำนวนเงิน', 'Amount': 'จำนวนเงิน'
        }
        for old_name, new_name in column_mapping.items():
            if old_name in pdf_df.columns:
                pdf_df = pdf_df.rename(columns={old_name: new_name})
        
        fchn_df = pd.read_excel(fchn_file, sheet_name=0, dtype=str)
        master_df = pd.read_excel(master_file, sheet_name=0, dtype=str)

        total_rows = len(pdf_df)
        
        # ==========================================
        # 1. จัดการ SHEET: TR Teams
        # ==========================================
        start_row_tr = 11
        st.info(f"📊 เริ่มประมวลผล TR Teams ({total_rows} แถว)")

        # --- เขียนข้อมูลลงแถวใหม่ (เฉพาะคอลัมน์ที่กำหนด) ---
        for idx, pdf_row in pdf_df.iterrows():
            row_num = start_row_tr + idx
            
            try:
                cheque_number = str(pdf_row['หมายเลขเช็ค'])
                amount = pdf_row['จำนวนเงิน']
                account_number = str(pdf_row['เลขบัญชี'])
                
                # Logic เดิมในการหา Business Partner และ Lookup ค่าต่างๆ
                if business_partner:
                    bp = business_partner
                else:
                    bp = xlookup(account_number, master_df.iloc[:, 4], master_df.iloc[:, 6])
                
                if bp: template_sheet.cell(row_num, 2).value = str(bp)
                
                # เขียนค่าลง Cell (จะไม่กระทบคอลัมน์อื่นที่เราไม่ได้สั่งแก้)
                template_sheet.cell(row_num, 6).value = "23.12.2025"
                template_sheet.cell(row_num, 10).value = "23.12.2025"
                template_sheet.cell(row_num, 8).value = amount
                template_sheet.cell(row_num, 15).value = f"CHQ{cheque_number}"
                template_sheet.cell(row_num, 31).value = str(account_number)
                
                # Lookup Logic (FCHN & Master) ...
                cheque_str = str(cheque_number)
                cheque_last8 = int(cheque_str[-8:]) if len(cheque_str) >= 8 else int(cheque_str)
                p_result = xlookup(cheque_last8, fchn_df.iloc[:, 0], fchn_df.iloc[:, 5])
                if p_result: template_sheet.cell(row_num, 16).value = str(p_result)
                
                if bp:
                    lookup_key = str(bp) + str(account_number)
                    i_result = xlookup(lookup_key, master_df.iloc[:, 12], master_df.iloc[:, 11])
                    if i_result: template_sheet.cell(row_num, 9).value = str(i_result)
                    
                    k_result = xlookup(lookup_key, master_df.iloc[:, 12], master_df.iloc[:, 7])
                    if k_result: 
                        template_sheet.cell(row_num, 11).value = str(k_result)
                        template_sheet.cell(row_num, 17).value = str(k_result)
                    
                    y_result = xlookup(lookup_key, master_df.iloc[:, 12], master_df.iloc[:, 8])
                    if y_result: 
                        template_sheet.cell(row_num, 25).value = str(y_result)
                        template_sheet.cell(row_num, 37).value = str(y_result)
                    
                    ac_result = xlookup(lookup_key, master_df.iloc[:, 12], master_df.iloc[:, 10])
                    if ac_result: template_sheet.cell(row_num, 29).value = str(ac_result)
                
                a_result = xlookup(account_number, master_df.iloc[:, 4], master_df.iloc[:, 0])
                if a_result: template_sheet.cell(row_num, 1).value = str(a_result)
                
                r_result = xlookup(account_number, master_df.iloc[:, 4], master_df.iloc[:, 9])
                if r_result: template_sheet.cell(row_num, 18).value = str(r_result)
                
                s_result = xlookup(account_number, master_df.iloc[:, 4], master_df.iloc[:, 1])
                if s_result: template_sheet.cell(row_num, 19).value = str(s_result).zfill(4)

            except Exception as e:
                st.error(f"❌ Error TR Row {idx+1}: {e}")
                continue

        st.success("✅ TR Teams: เสร็จสมบูรณ์")

        # ==========================================
        # 2. จัดการ SHEET: Cash Teams
        # ==========================================
        start_row_cash = 6
        st.info(f"📊 เริ่มประมวลผล Cash Teams")

        # --- Phase 1: เขียนข้อมูลลงแถวใหม่ ---
        for idx, pdf_row in pdf_df.iterrows():
            cash_row = start_row_cash + idx
            try:
                cheque_number = str(pdf_row['หมายเลขเช็ค'])
                amount = pdf_row['จำนวนเงิน']
                account_number = str(pdf_row['เลขบัญชี'])
                
                if business_partner:
                    bp = business_partner
                else:
                    bp = xlookup(account_number, master_df.iloc[:, 4], master_df.iloc[:, 6])
                
                company_code = xlookup(account_number, master_df.iloc[:, 4], master_df.iloc[:, 0])
                if company_code: cash_sheet.cell(cash_row, 1).value = str(company_code)
                
                business_place = xlookup(account_number, master_df.iloc[:, 4], master_df.iloc[:, 1])
                if business_place: cash_sheet.cell(cash_row, 2).value = str(business_place).zfill(4)
                
                cash_sheet.cell(cash_row, 5).value = "23.12.2025"
                cash_sheet.cell(cash_row, 6).value = amount
                cash_sheet.cell(cash_row, 7).value = str(account_number)
                
                company_name = xlookup(str(account_number), fchn_df.iloc[:, 8], fchn_df.iloc[:, 7])
                if company_name and str(company_name).lower() not in ['none', 'nan', '']:
                    cash_sheet.cell(cash_row, 3).value = str(company_name)
                
                house_bank = xlookup(str(account_number), fchn_df.iloc[:, 8], fchn_df.iloc[:, 2])
                if house_bank and str(house_bank).lower() not in ['none', 'nan', '']:
                    bank_name_only = re.sub(r'\d+', '', str(house_bank)).strip()
                    cash_sheet.cell(cash_row, 4).value = bank_name_only
                
                cash_sheet.cell(cash_row, 8).value = f"CHQ{cheque_number}"
                if bp: cash_sheet.cell(cash_row, 9).value = str(bp)

            except Exception as e:
                st.error(f"❌ Error Cash Row {idx+1}: {e}")
                continue

        st.success("✅ Cash Teams: เสร็จสมบูรณ์")

        # Save
        output = BytesIO()
        template_wb.save(output)
        template_wb.close()
        output.seek(0)
        
        return output, len(pdf_df)
        
    except Exception as e:
        st.error(f"❌ เกิดข้อผิดพลาด: {str(e)}")
        st.code(traceback.format_exc())
        return None, 0

# ========== Main UI ==========
def main():
    st.title("🏦 ระบบดึงข้อความจากเช็คไทย")
    st.caption("📌 ใช้ EasyOCR (ไทย + อังกฤษ) + Tesseract MICR (e13b)")
    
    tab1, tab2 = st.tabs(["📄 ดึงข้อความจากเช็ค", "📊 เติมข้อมูล Template"])
    
    # ===== TAB 1: OCR Extraction =====
    with tab1:
        st.markdown("### 📤 อัปโหลดไฟล์เช็ค")
        uploaded_file = st.file_uploader(
            "เลือกไฟล์ PDF หรือ รูปภาพ",
            type=['pdf', 'jpg', 'jpeg', 'png'],
            help="รองรับไฟล์ PDF และรูปภาพ (JPG, PNG)",
            key="ocr_upload"
        )
        
        if uploaded_file:
            col1, col2 = st.columns([1, 1])
            
            with col1:
                try:
                    uploaded_file.seek(0)
                    if uploaded_file.name.lower().endswith('.pdf'):
                        images = convert_from_bytes(uploaded_file.read(), dpi=150)
                        st.image(images[0], caption="ไฟล์ที่อัปโหลด", use_container_width=True)
                    else:
                        st.image(uploaded_file, caption="ไฟล์ที่อัปโหลด", use_container_width=True)
                except Exception as e:
                    st.warning(f"⚠️ ไม่สามารถแสดงตัวอย่าง: {str(e)}")
            
            with col2:
                if st.button("🚀 เริ่มประมวลผล", type="primary", use_container_width=True):
                    with st.spinner("⏳ กำลังประมวลผล..."):
                        all_results, all_cropped = process_cheque(uploaded_file)
                        
                        if all_results:
                            st.success(f"✅ ประมวลผลสำเร็จ! พบ {len(all_results)} หน้า")
                            
                            df_result = pd.DataFrame(all_results)
                            st.dataframe(df_result, use_container_width=True)
                            
                            # Download Excel
                            output = BytesIO()
                            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                                df_result.to_excel(writer, sheet_name='Cheques', index=False)
                                
                                workbook = writer.book
                                worksheet = writer.sheets['Cheques']
                                text_format = workbook.add_format({'num_format': '@'})
                                
                                for col_name in ["หมายเลขเช็ค", "เลขบัญชี", "Cheque digit", "รหัสธนาคาร", "รหัสสาขา"]:
                                    if col_name in df_result.columns:
                                        col_idx = df_result.columns.get_loc(col_name)
                                        worksheet.set_column(col_idx, col_idx, 20, text_format)
                            
                            output.seek(0)
                            st.download_button(
                                label="📥 ดาวน์โหลด Excel",
                                data=output,
                                file_name="cheque_data.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                            
                            with st.expander("🖼️ ดูภาพที่ Crop แล้ว"):
                                for idx, cropped in enumerate(all_cropped, start=1):
                                    st.image(
                                        cv2.cvtColor(cropped, cv2.COLOR_BGR2RGB), 
                                        caption=f"หน้า {idx}",
                                        use_container_width=True
                                    )
    
    # ===== TAB 2: Template Filling =====
    with tab2:
        st.markdown("### 📤 อัปโหลดไฟล์ที่จำเป็น")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**📂 ไฟล์หลัก**")
            
            template_file = st.file_uploader(
                "1️⃣ Template File",
                type=['xlsx'],
                help="Template TR & Cash.xlsx",
                key="template_uploader"
            )
            
            pdf_file = st.file_uploader(
                "2️⃣ Extracted Data (จาก Tab 1)",
                type=['xlsx'],
                help="ไฟล์ Excel ที่ดึงข้อมูลจากเช็คแล้ว",
                key="pdf_uploader"
            )
        
        with col2:
            st.markdown("**📊 ไฟล์ Lookup**")
            
            fchn_file = st.file_uploader(
                "3️⃣ FCHN File",
                type=['xlsx'],
                help="FCHN.xlsx สำหรับ Lookup",
                key="fchn_uploader"
            )
            
            master_file = st.file_uploader(
                "4️⃣ Master File",
                type=['xlsx'],
                help="Copy of Master File*.xlsx",
                key="master_uploader"
            )
        
        st.markdown("---")
        st.markdown("### ⚙️ การตั้งค่า")
        business_partner = st.text_input(
            "Business Partner (ถ้าไม่ระบุจะใช้ Auto-lookup)",
            placeholder="เช่น UOB0052, CIM0199, TNB0497",
            help="ปล่อยว่างไว้เพื่อให้ระบบค้นหาจาก Master file อัตโนมัติ"
        )
        
        st.markdown("---")
        
        if st.button(
            "🚀 ประมวลผล Template", 
            type="primary", 
            use_container_width=True,
            disabled=not all([template_file, pdf_file, fchn_file, master_file])
        ):
            if all([template_file, pdf_file, fchn_file, master_file]):
                try:
                    with st.spinner("⏳ กำลังประมวลผล Template..."):
                        output, total_rows = process_template_filling(
                            pdf_file,
                            fchn_file,
                            master_file,
                            template_file,
                            business_partner.strip()
                        )
                        
                        if output:
                            st.success(f"✅ ประมวลผลสำเร็จ! จำนวน {total_rows} แถว")
                            st.info(f"""
                            📋 สรุปผลการประมวลผล:
                            - จำนวนแถวที่ประมวลผล: {total_rows} แถว
                            - TR Teams: แถว 11 ถึง {11 + total_rows - 1}
                            - Cash Teams: แถว 6 ถึง {6 + total_rows - 1}
                            """)
                            
                            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                            st.download_button(
                                label="📥 ดาวน์โหลด Template ที่เติมข้อมูลแล้ว",
                                data=output,
                                file_name=f"Template_PDF_Filled_{timestamp}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                
                except Exception as e:
                    st.error(f"❌ เกิดข้อผิดพลาด: {str(e)}")
                    st.code(traceback.format_exc())
            else:
                st.warning("⚠️ กรุณาอัปโหลดไฟล์ให้ครบทั้ง 4 ไฟล์!")
        
        st.markdown("---")
        st.markdown("""
        ### 📝 วิธีใช้งาน
        1. อัปโหลด **Template TR & Cash.xlsx**
        2. อัปโหลด **Extracted Data** จาก Tab 1
        3. อัปโหลด **FCHN.xlsx** และ **Master File**
        4. กดปุ่ม **ประมวลผล Template**
        5. ดาวน์โหลดไฟล์ที่เติมข้อมูลแล้ว
        
        ### ℹ️ ระบบจะทำอะไรบ้าง
        - เติมข้อมูลลง TR Teams & Cash Teams sheets
        - ทำ XLOOKUP จาก FCHN & Master อัตโนมัติ
        - พร้อม Import เข้า SAP ได้เลย
        """)

if __name__ == "__main__":
    main()
